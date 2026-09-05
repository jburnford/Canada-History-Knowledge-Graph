#!/usr/bin/env python3
"""Stage the canonical census workbooks without discarding source cells.

Select structured worksheets by their headers, not workbook position. Keep CD
totals, table-scoped identifiers, repeated columns, and unresolved semantics.
This is source preservation; geographic identity requires separate evidence.
"""

import argparse
import hashlib
import json
import os
import re
import sqlite3
import tempfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from _config import CONFIG, REPO_ROOT
from stage_1911_reporting_units import cell_value, parse_dls, reporting_level

META = {'YEAR', 'PR', 'ROW_ID', 'NOTES', 'CD_NO', 'CSD_NO', 'LINE_NO',
        'PR_CD', 'PR_CD_CSD', 'CSD_TYPE'}
PERSON_TABLES = {'1911_V1T2', '1911_V2T2', '1911_V2T7', '1911_V2T28',
                 '1921_V1T16', '1921_V1T27', '1921_V1T38'}


def discover(root):
    sources = []
    for year in range(1851, 1922, 10):
        folder = root / (f'{year}Tables/{year}' if year >= 1911 else str(year))
        patterns = [f'{year}_*_PUB_*.xlsx'] if year >= 1911 else [
            f'{year}_*_CD_*.xlsx', f'{year}_*_CSD_*.xlsx']
        sources.extend((year, path) for pattern in patterns for path in folder.glob(pattern))
    return sorted(sources)


def is_metadata(header):
    return header in META or bool(re.fullmatch(
        r'(?:NAME_(?:CD|CSD|COUNTY)|NUMBER_(?:CD|CSD)|TCPUID_(?:CD|CSD)|V\d+T\d+)_\d{4}', header))


def structured_sheet(workbook):
    candidates = []
    for sheet in workbook:
        for row in sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row), values_only=True):
            headers = {str(value).strip() for value in row if value is not None}
            if 'PR' in headers and any(h.startswith('TCPUID_') or re.fullmatch(r'V\d+T\d+_\d{4}', h) for h in headers):
                candidates.append((sheet, row))
                break
    if len(candidates) != 1:
        raise ValueError(f'Expected one structured worksheet; found {[s.title for s, _ in candidates]}')
    sheet, header = candidates[0]
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row == header:
            return sheet, index, header
    raise AssertionError('Selected header disappeared')


def master_definitions(path):
    workbook = load_workbook(path, read_only=True, data_only=False)
    definitions = {}
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    sheet = workbook.active
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:
            continue
        key = str(row[0]).strip()
        value = dict(description=str(row[1] or ''), category=str(row[2] or ''),
                     source_path=str(path), source_sha256=digest,
                     source_cell=f'{sheet.title}!B{index}', raw_row=list(row))
        if key in definitions:
            if any(definitions[key][field] != value[field] for field in ['description', 'category']):
                raise ValueError(f'Conflicting master definitions: {key}')
            definitions[key].setdefault('additional_source_rows', []).append(value)
        else:
            definitions[key] = value
    workbook.close()
    return definitions


def semantics(header, definition, table_key, vintage):
    """Conservative units and periods; unspecified historical units stay unknown."""
    description = definition.get('description', '')
    lower = description.casefold()
    base = re.sub(r'_\d{4}$', '', header)
    unit, unit_status = None, 'requires_source_interpretation'
    # Order matters: population density, family size and currency are not
    # population counts or pounds of mass. No conversion to modern SI/CAD.
    if 'PER_SQ_MI' in header:
        unit = 'persons per square mile'
    elif base == 'POP_AVF_PR':
        unit = 'persons per family'
    elif base == 'DTH_BR_PR':
        unit = 'deaths per birth'
    elif header.startswith('PCT_') or 'percentage' in lower:
        unit = 'percent'
    elif 'pounds sterling' in lower:
        unit = 'pounds sterling'
    elif 'dollars' in lower:
        unit = 'dollars as reported'
    elif header.endswith('_N') and description:
        # A count of the entities described by this variable, not an assertion
        # that houses, animals or families are persons.
        unit = 'count'
    else:
        for pattern, name in [
            (r'\bacres?\b', 'acres'), (r'\bsquare miles?\b', 'square miles'),
            (r'\bbushels?\b', 'bushels as reported'), (r'\bbarrels?\b', 'barrels as reported'),
            (r'\bpounds?\b', 'pounds of mass as reported'), (r'\btons?\b', 'tons as reported'),
            (r'\byards?\b', 'yards'), (r'\bgallons?\b', 'gallons as reported'),
            (r'\bquintals?\b', 'quintals as reported'), (r'\bfathoms?\b', 'fathoms'),
            (r'\bfeet of lumber\b', 'feet of lumber as reported'),
            (r'^logs produced\b', 'logs')]:
            if re.search(pattern, lower):
                unit = name
                break
    if unit:
        unit_status = 'source_definition' if definition else 'explicit_column_semantics'
    if not unit and definition and re.search(r'(?:^|_)DWELLINGS(?:_|$)', header):
        unit, unit_status = 'dwellings', 'source_column_and_definition'
    if not unit and definition and re.search(r'(?:^|_)HOUSEHOLDS(?:_|$)', header):
        unit, unit_status = 'households', 'source_column_and_definition'
    if not unit and definition and base == 'FAMILIES':
        unit, unit_status = 'families', 'source_column_and_definition'
    if not unit and definition and (table_key in PERSON_TABLES or 'POP' in base.split('_')):
        unit, unit_status = 'persons', 'population_table_context_and_definition'

    explicit_year = re.search(r'_(\d{4})$', header)
    year = int(explicit_year.group(1)) if explicit_year else None
    period = 'explicit_column_year' if year else 'reference_period_not_resolved'
    if re.search(r'past year|last year|preceding year', lower):
        year, period = None, 'preceding_year_as_described'
    elif re.search(r'per day|daily', lower):
        year, period = None, 'daily_production_as_described'
    elif re.search(r'per week|weekly', lower):
        year, period = None, 'weekly_production_as_described'
    elif re.search(r'per year|annual', lower):
        year, period = None, 'annual_production_as_described'
    return dict(variable=base, reference_year=year, unit=unit, unit_status=unit_status,
                reference_period_kind=period, reference_period_text=description,
                definition=description, definition_source_json=json.dumps(definition, ensure_ascii=False))


SCHEMA = '''
CREATE TABLE source (path TEXT, sha256 TEXT, sheet TEXT, census_vintage INTEGER,
 source_key TEXT, table_key TEXT, header_row INTEGER, worksheets_json TEXT);
CREATE TABLE reporting_units (
 unit_id TEXT PRIMARY KEY, source_code TEXT, excel_row INTEGER UNIQUE,
 label TEXT, province TEXT, reporting_level TEXT, survey_unit_id TEXT,
 candidate_snapshot_id TEXT, spatial_binding_status TEXT,
 legacy_same_code_gis_name TEXT, code_collision INTEGER, source_metadata_json TEXT NOT NULL);
CREATE TABLE observations (
 observation_id TEXT PRIMARY KEY, unit_id TEXT NOT NULL REFERENCES reporting_units(unit_id),
 source_cell TEXT UNIQUE, source_column TEXT, variable TEXT, reference_year INTEGER,
 reporting_geography_vintage INTEGER, unit TEXT, raw_value_json TEXT,
 numeric_value TEXT, value_status TEXT, excel_number_format TEXT,
 column_key TEXT, unit_status TEXT, reference_period_kind TEXT, reference_period_text TEXT);
CREATE TABLE survey_units (
 survey_unit_id TEXT PRIMARY KEY, township INTEGER, range_number INTEGER,
 meridian_direction TEXT, meridian INTEGER, geometry_status TEXT);
CREATE TABLE source_columns (
 column_key TEXT PRIMARY KEY, column_index INTEGER UNIQUE, source_column TEXT,
 column_role TEXT, definition TEXT, definition_source_json TEXT,
 variable TEXT, reference_year INTEGER, unit TEXT, unit_status TEXT,
 reference_period_kind TEXT, reference_period_text TEXT);
CREATE TABLE source_rows (excel_row INTEGER PRIMARY KEY, raw_values_json TEXT);
CREATE TABLE definition_rows (source_row INTEGER PRIMARY KEY, raw_values_json TEXT);
'''


def stage(workbook_path, vintage, out, master):
    out.mkdir(parents=True, exist_ok=True)
    source_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    source_key = workbook_path.stem
    match = re.match(r'(\d{4}_(?:PE_)?V\d+T\d+)', source_key)
    if not match:
        raise ValueError(f'Unrecognized source filename: {source_key}')
    table_key = match.group(1)
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet, header_row, raw_headers = structured_sheet(workbook)
    # One bounded workbook in memory, retaining all physical column positions,
    # including unnamed columns with values. Empty formatted columns are inert.
    rows = [(index, row) for index, row in enumerate(sheet.iter_rows(min_row=header_row+1), header_row+1)
            if any(cell.value is not None for cell in row)]
    active = [i for i, h in enumerate(raw_headers) if h is not None or any(row[i].value is not None for _, row in rows)]
    headers = {i: str(raw_headers[i]).strip() if raw_headers[i] is not None else '' for i in active}
    meta_headers = [h for h in headers.values() if is_metadata(h)]
    if len(meta_headers) != len(set(meta_headers)):
        raise ValueError('Ambiguous duplicate metadata columns')
    definitions, definition_rows = dict(master), []
    if 'Variables' in workbook.sheetnames:
        for index, row in enumerate(workbook['Variables'].iter_rows(values_only=True), 1):
            definition_rows.append((index, json.dumps(list(row), ensure_ascii=False)))
            if row[0] is None:
                continue
            key = str(row[0]).strip()
            if key in definitions and definitions[key].get('source_sha256') == source_hash:
                raise ValueError(f'Duplicate local variable definition: {key}')
            definitions[key] = dict(description=str(row[1] or ''), source_path=str(workbook_path),
                                    source_sha256=source_hash, source_cell=f'Variables!B{index}', raw_row=list(row))
    columns = {}
    for i, header in headers.items():
        role = 'metadata' if is_metadata(header) else 'statistical' if header else 'unlabelled_source_column'
        columns[i] = dict(semantics(header, definitions.get(header, {}), table_key, vintage),
                          column_key=get_column_letter(i+1), column_index=i+1, source_column=header, column_role=role)
    fd, temporary = tempfile.mkstemp(prefix='census-', suffix='.sqlite', dir=out)
    os.close(fd)
    db = sqlite3.connect(temporary)
    try:
        db.executescript(SCHEMA)
        db.execute('PRAGMA foreign_keys = ON')
        db.execute('INSERT INTO source VALUES (?,?,?,?,?,?,?,?)',
                   (str(workbook_path), source_hash, sheet.title, vintage, source_key, table_key,
                    header_row, json.dumps(workbook.sheetnames)))
        db.executemany('INSERT INTO definition_rows VALUES (?,?)', definition_rows)
        for col in columns.values():
            db.execute('INSERT INTO source_columns VALUES (:column_key,:column_index,:source_column,:column_role,'
                       ':definition,:definition_source_json,:variable,:reference_year,:unit,:unit_status,'
                       ':reference_period_kind,:reference_period_text)', col)
        levels, statuses, codes, unit_statuses, periods = Counter(), Counter(), Counter(), Counter(), Counter()
        for row_no, row in rows:
            meta = {headers[i]: row[i].value for i in active if columns[i]['column_role'] == 'metadata'}
            name = str(meta.get(f'NAME_CSD_{vintage}') or meta.get('PR_CD_CSD') or
                       meta.get(f'NAME_CD_{vintage}') or meta.get(f'NAME_COUNTY_{vintage}') or meta.get('PR_CD') or '')
            code = str(meta.get(f'TCPUID_CSD_{vintage}') or meta.get(f'TCPUID_CD_{vintage}') or
                       meta.get(table_key.split('_')[-1] + f'_{vintage}') or '')
            codes[code] += 1
            # Source rows are distinct statistical subjects even if identifiers
            # repeat. No cross-table or same-code GIS identity is asserted.
            unit_id = f'REPORTING_{source_key}_ROW_{row_no}'
            dls = parse_dls(name)
            if f'TCPUID_CSD_{vintage}' in meta:
                level = 'census_subdivision_reporting_unit'
            elif f'TCPUID_CD_{vintage}' in meta:
                level = 'census_division_total'
            else:
                level = reporting_level(meta, dls)
                if level == 'unresolved_reporting_level' and 'PR_CD' in meta and meta.get('CD_NO') is not None:
                    level = 'census_division_total'
            if dls:
                level = 'survey_township_reporting_unit'
                db.execute('INSERT OR IGNORE INTO survey_units VALUES (?,?,?,?,?,?)',
                           (dls['survey_unit_id'], dls['township'], dls['range'],
                            dls['meridian_direction'], dls['meridian'], 'not_supplied'))
            binding = 'survey_identifier_only_geometry_unresolved' if dls else 'no_validated_spatial_binding'
            db.execute('INSERT INTO reporting_units VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                       (unit_id, code, row_no, name, meta.get('PR'), level,
                        dls['survey_unit_id'] if dls else '', '', binding, '', 0,
                        json.dumps(meta, ensure_ascii=False)))
            db.execute('INSERT INTO source_rows VALUES (?,?)',
                       (row_no, json.dumps({get_column_letter(i+1): row[i].value for i in active}, ensure_ascii=False)))
            levels[level] += 1
            for i, col in columns.items():
                if col['column_role'] == 'metadata':
                    continue
                cell = row[i]
                numeric, status = cell_value(cell.value, cell.data_type)
                db.execute('INSERT INTO observations VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                           (f'CELL_{source_key}_{row_no}_{i+1}', unit_id, f'{sheet.title}!{col["column_key"]}{row_no}',
                            col['source_column'], col['variable'], col['reference_year'], vintage, col['unit'],
                            json.dumps(cell.value, ensure_ascii=False), numeric, status, cell.number_format or 'General',
                            col['column_key'], col['unit_status'], col['reference_period_kind'], col['reference_period_text']))
                statuses[status] += 1; unit_statuses[col['unit_status']] += 1; periods[col['reference_period_kind']] += 1
        db.commit()
        cell_count = db.execute('SELECT COUNT(*) FROM observations').fetchone()[0]
        expected = len(rows) * sum(c['column_role'] != 'metadata' for c in columns.values())
        errors = db.execute('PRAGMA foreign_key_check').fetchall()
        if expected != cell_count or errors:
            raise ValueError('Source-cell count or foreign key reconciliation failed')
        manifest = dict(source=str(workbook_path), source_sha256=source_hash, source_key=source_key,
                        census_vintage=vintage, selected_sheet=sheet.title, header_row=header_row,
                        other_sheets=[s for s in workbook.sheetnames if s != sheet.title],
                        reporting_rows=len(rows), preserved_cells=cell_count, expected_cells=expected,
                        reporting_levels=dict(levels), value_statuses=dict(statuses),
                        unit_status_cells=dict(unit_statuses), reference_period_cells=dict(periods),
                        duplicate_source_codes={k: v for k, v in codes.items() if k and v > 1},
                        duplicate_statistical_headers={k: v for k, v in Counter(headers.values()).items() if k and v > 1},
                        ignored_empty_unnamed_columns=len(raw_headers)-len(active),
                        unlabelled_nonempty_columns=[c['column_key'] for c in columns.values() if c['column_role'].startswith('unlabelled')],
                        source_scope='Canonical structured worksheets; alternative OCR sheets are not independent observations',
                        status='Source preservation; geographic bindings and unresolved semantics require separate reconciliation')
    finally:
        db.close()
        workbook.close()
    os.replace(temporary, out / 'source_observations.sqlite')
    (out / 'manifest.json').write_text(json.dumps(manifest, indent=2, ensure_ascii=False)+'\n')
    return manifest


def validate_source(database):
    """Independently reread source coordinates, values, cell types and formats."""
    db = sqlite3.connect(database)
    db.row_factory = sqlite3.Row
    source = db.execute('SELECT * FROM source').fetchone()
    path = Path(source['path'])
    errors = []
    if hashlib.sha256(path.read_bytes()).hexdigest() != source['sha256']:
        errors.append('Source workbook hash changed')
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook[source['sheet']]
    columns = list(db.execute('SELECT * FROM source_columns ORDER BY column_index'))
    seen_rows, cells = set(), 0
    for row_no, row in enumerate(sheet.iter_rows(min_row=source['header_row']+1), source['header_row']+1):
        if all(cell.value is None for cell in row):
            continue
        seen_rows.add(row_no)
        raw = db.execute('SELECT raw_values_json FROM source_rows WHERE excel_row=?', (row_no,)).fetchone()
        expected = {c['column_key']: row[c['column_index']-1].value for c in columns}
        if raw is None or json.loads(raw[0]) != expected:
            errors.append(f'Row {row_no}: source metadata/value mismatch')
        for col in columns:
            if col['column_role'] == 'metadata':
                continue
            cell = row[col['column_index']-1]
            coordinate = f'{sheet.title}!{col["column_key"]}{row_no}'
            stored = db.execute('SELECT raw_value_json,numeric_value,value_status,excel_number_format FROM observations WHERE source_cell=?',
                                (coordinate,)).fetchone()
            numeric, status = cell_value(cell.value, cell.data_type)
            if stored is None or tuple(stored) != (json.dumps(cell.value, ensure_ascii=False), numeric, status, cell.number_format or 'General'):
                errors.append(f'{coordinate}: source statistical cell mismatch')
            cells += 1
    if seen_rows != {r[0] for r in db.execute('SELECT excel_row FROM reporting_units')}:
        errors.append('Reporting row sets differ')
    if cells != db.execute('SELECT COUNT(*) FROM observations').fetchone()[0]:
        errors.append('Statistical cell counts differ')
    workbook.close(); db.close()
    return dict(source_rows=len(seen_rows), source_cells=cells, errors=errors)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--data-root', type=Path, default=CONFIG.data_root)
    ap.add_argument('--out', type=Path, default=REPO_ROOT/'data_quality/lod_census_sources')
    ap.add_argument('--only', help='Only filenames containing this string')
    args = ap.parse_args()
    master = master_definitions(args.data_root/'1911Tables/1911/TCP_CANADA_CD-CSD_Mastvar.xlsx')
    results = []
    for year, path in discover(args.data_root):
        if args.only and args.only not in path.name:
            continue
        directory = args.out/path.stem
        result = stage(path, year, directory, master)
        validation = validate_source(directory/'source_observations.sqlite')
        (directory/'source.validation.json').write_text(json.dumps(validation, indent=2)+'\n')
        if validation['errors']:
            raise ValueError(validation)
        results.append(result)
        print(json.dumps(dict(source=path.name, rows=result['reporting_rows'], cells=result['preserved_cells'],
                              validation='passed')), flush=True)
    if not results:
        raise ValueError('No canonical workbooks selected')
    catalog = dict(workbooks=len(results), reporting_rows=sum(r['reporting_rows'] for r in results),
                   preserved_cells=sum(r['preserved_cells'] for r in results), sources=results)
    name = 'catalog.json' if not args.only else 'selected-catalog.json'
    (args.out/name).write_text(json.dumps(catalog, indent=2, ensure_ascii=False)+'\n')


if __name__ == '__main__':
    main()

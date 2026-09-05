#!/usr/bin/env python3
"""Assess source reporting units against the audited census map inventory.

Explicit older TCPUID fields require matching year, province, district and
name. PUB codes are never presumed to be GIS identifiers. Name candidates
require a district supplied by their own table. No fuzzy name search, survey
geometry inference, historical identity merge or population apportionment.
"""

import argparse
import csv
import hashlib
import json
import re
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from _config import REPO_ROOT
from _normalize import normalize_for_match


def name_key(value):
    value = re.sub(r'[\u2010-\u2015\u2212]', ' ', value or '')
    return re.sub(r'\s+', ' ', re.sub(r'[.,]', ' ', normalize_for_match(value))).strip()


def district_number(value):
    if value is None:
        return ''
    return str(value).strip().removesuffix('.0')


def source_names(label, metadata):
    # Retain administrative tiers. These spelling variants do not turn a
    # township into the same-name town, village or city.
    names = {name_key(label)}
    kind = str(metadata.get('CSD_TYPE') or '').strip().upper()
    suffixes = {'C': ['C', 'City'], 'T': ['T', 'Town'], 'VL': ['VL', 'Village'],
                'TP': ['TP', 'Township'], 'PAR': ['PAR', 'Parish'],
                'R': ['R', 'Reserve']}.get(kind, [])
    if suffixes:
        names = {name_key(label + ' ' + suffix) for suffix in suffixes}
        # Some labels already contain the explicit tier from CSD_TYPE.
        if any(name_key(label).endswith(' ' + name_key(s)) for s in suffixes):
            names.add(name_key(label))
    return names


class Inventory:
    def __init__(self, representations):
        self.by_id = {}
        self.by_name = defaultdict(list)
        for row in representations:
            key = (row['source_id'], int(row['year']))
            if key in self.by_id:
                raise ValueError(f'Duplicate map representation: {key}')
            self.by_id[key] = row
            if str(row['is_coverage_record']).lower() == 'true':
                continue
            district = name_key(row.get('cd_name', '')) if row['level'] == 'csd' else ''
            name = name_key(row['name'])
            self.by_name[(row['level'], int(row['year']), row['province'].strip().upper(), district, name)].append(row['snapshot_id'])

    def resolve(self, row, vintage, parents, cached_metadata=None):
        metadata = json.loads(row['source_metadata_json'])
        metadata.update(cached_metadata or {})
        level = row['reporting_level']
        province = str(metadata.get('PR', row['province']) or '').strip().upper()
        code = str(metadata.get(f'TCPUID_CSD_{vintage}') or metadata.get(f'TCPUID_CD_{vintage}') or row['source_code']).strip()
        same_code = self.by_id.get((code, vintage))
        result = dict(snapshot_ids=[], status='no_contextual_name_match',
                      same_code_snapshot=same_code['snapshot_id'] if same_code else '',
                      same_code_gis_name=same_code['name'] if same_code else '',
                      district='', method='none', effective_source_code=code, effective_province=province)
        if row['survey_unit_id']:
            return dict(result, status='survey_identifier_only_geometry_unresolved')
        if level in {'country_total', 'province_or_territory_total'}:
            return dict(result, status='aggregate_outside_csd_cd_map_inventory')
        if level == 'unresolved_reporting_level':
            return dict(result, status='reporting_level_unresolved')
        if level not in {'census_division_total', 'census_subdivision_reporting_unit'}:
            return dict(result, status='reporting_level_requires_interpretation')
        if any(isinstance(value, str) and value.startswith('=') for key, value in metadata.items()
               if key not in {'NOTES', 'YEAR', 'ROW_ID'}):
            return dict(result, status='source_metadata_formula_without_cached_result')
        original = json.loads(row['source_metadata_json'])
        if any(value is None and isinstance(original.get(key), str) and original[key].startswith('=')
               for key, value in (cached_metadata or {}).items() if key not in {'NOTES', 'YEAR', 'ROW_ID'}):
            return dict(result, status='source_metadata_formula_without_cached_result')
        map_level = 'cd' if level == 'census_division_total' else 'csd'
        district = metadata.get(f'NAME_CD_{vintage}') or metadata.get(f'NAME_COUNTY_{vintage}')
        if map_level == 'csd' and not district:
            choices = parents.get((province, district_number(metadata.get('CD_NO'))), set())
            if len(choices) != 1:
                return dict(result, status='parent_district_missing_or_ambiguous')
            district = next(iter(choices))
        result['district'] = district or ''
        names = source_names(row['label'], metadata) if map_level == 'csd' else {name_key(row['label'])}
        explicit_code = metadata.get(f'TCPUID_CSD_{vintage}') if map_level == 'csd' else None
        if explicit_code:
            if same_code is None:
                return dict(result, status='explicit_source_identifier_missing_from_map', method='explicit_tcpuid')
            if str(same_code['is_coverage_record']).lower() == 'true':
                return dict(result, status='map_identifier_is_coverage_record', method='explicit_tcpuid')
            if (same_code['province'] != province or same_code['level'] != map_level or
                    name_key(same_code['name']) not in names or
                    name_key(same_code.get('cd_name', '')) != name_key(district)):
                return dict(result, status='explicit_identifier_context_conflict', method='explicit_tcpuid')
            return dict(result, snapshot_ids=[same_code['snapshot_id']],
                        status='source_identifier_and_context_agree', method='explicit_tcpuid_name_province_district_year')
        candidates = set()
        for name in names:
            key = (map_level, vintage, province, name_key(district) if map_level == 'csd' else '', name)
            candidates.update(self.by_name.get(key, []))
        return dict(result, snapshot_ids=sorted(candidates), method='name_province_district_year',
                    status='unique_contextual_name_candidate' if len(candidates) == 1 else
                    'ambiguous_contextual_name_candidates' if candidates else 'no_contextual_name_match')


def cached_metadata(db, rows):
    """Read stored workbook results, without executing or repairing formulas."""
    needed = {}
    for row in rows:
        fields = {key for key, value in json.loads(row['source_metadata_json']).items()
                  if isinstance(value, str) and value.startswith('=')}
        if fields:
            needed[row['excel_row']] = fields
    if not needed:
        return {}
    source = db.execute('SELECT * FROM source').fetchone()
    path = Path(source['path'])
    if hashlib.sha256(path.read_bytes()).hexdigest() != source['sha256']:
        raise ValueError('Source workbook changed since observation staging')
    columns = {row['source_column']: row['column_index']-1 for row in db.execute(
        "SELECT * FROM source_columns WHERE column_role='metadata'")}
    workbook = load_workbook(path, read_only=True, data_only=True)
    result = {}
    for row_no, row in enumerate(workbook[source['sheet']].iter_rows(values_only=True), 1):
        if row_no in needed:
            result[row_no] = {field: row[columns[field]] for field in needed[row_no]}
    workbook.close()
    return result


def validate_bindings(root, representations, bindings):
    """Check every emitted source row and target against independent inventories."""
    catalog = json.loads((root/'catalog.json').read_text())
    expected = {}
    for source in catalog['sources']:
        db = sqlite3.connect(root/source['source_key']/'source_observations.sqlite')
        for unit_id, excel_row, level in db.execute('SELECT unit_id,excel_row,reporting_level FROM reporting_units'):
            expected[unit_id] = (source['source_key'], str(excel_row), level, str(source['census_vintage']))
        db.close()
    with representations.open() as stream:
        targets = {r['snapshot_id']: r for r in csv.DictReader(stream)}
    errors, seen, endpoints = [], set(), 0
    with bindings.open() as stream:
        for row in csv.DictReader(stream):
            uid = row['unit_id']
            if uid in seen:
                errors.append(f'Duplicate source reporting unit: {uid}')
            seen.add(uid)
            actual = (row['source'], row['excel_row'], row['reporting_level'], row['census_vintage'])
            if expected.get(uid) != actual:
                errors.append(f'Source row metadata mismatch: {uid}')
            candidates = json.loads(row['snapshot_ids_json'])
            endpoints += len(candidates)
            if len(set(candidates)) != len(candidates):
                errors.append(f'Duplicate candidate endpoint: {uid}')
            status = row['status']
            if status in {'source_identifier_and_context_agree','unique_contextual_name_candidate'}:
                if len(candidates) != 1:
                    errors.append(f'Expected exactly one endpoint: {uid}')
            elif status == 'ambiguous_contextual_name_candidates':
                if len(candidates) < 2:
                    errors.append(f'Expected multiple unresolved endpoints: {uid}')
            elif candidates:
                errors.append(f'Unresolved source row acquired a geographic binding: {uid}')
            for candidate in candidates:
                target = targets.get(candidate)
                if target is None:
                    errors.append(f'Missing map endpoint: {candidate}')
                    continue
                level = 'cd' if row['reporting_level'] == 'census_division_total' else 'csd'
                if target['year'] != row['census_vintage'] or target['level'] != level:
                    errors.append(f'Wrong census vintage or reporting level: {uid}')
                if str(target['is_coverage_record']).lower() == 'true':
                    errors.append(f'Coverage record acquired a population binding: {uid}')
    if seen != set(expected):
        errors.append('Source reporting unit sets differ')
    return dict(reporting_rows=len(seen), assessed_map_endpoints=endpoints, errors=errors)


def build(root, representations, out):
    with representations.open() as stream:
        inventory = Inventory(list(csv.DictReader(stream)))
    catalog = json.loads((root/'catalog.json').read_text())
    out.mkdir(parents=True, exist_ok=True)
    statuses = Counter()
    endpoint_count = 0
    cache_rows = 0
    source_hashes = {}
    context_reviews = []
    fields = ['source', 'unit_id', 'excel_row', 'source_code', 'label', 'province', 'census_vintage',
              'reporting_level', 'district', 'status', 'method', 'snapshot_ids_json',
              'same_code_snapshot', 'same_code_gis_name', 'effective_source_code', 'effective_province', 'cached_metadata_json']
    with (out/'bindings.csv').open('w', newline='') as stream:
        writer = csv.DictWriter(stream, fieldnames=fields); writer.writeheader()
        for source in catalog['sources']:
            key = source['source_key']
            database = root/key/'source_observations.sqlite'
            source_hashes[key] = hashlib.sha256(database.read_bytes()).hexdigest()
            db = sqlite3.connect(database); db.row_factory = sqlite3.Row
            rows = list(db.execute('SELECT * FROM reporting_units ORDER BY excel_row'))
            caches = cached_metadata(db, rows)
            parents = defaultdict(set)
            for row in rows:
                if row['reporting_level'] == 'census_division_total':
                    metadata = json.loads(row['source_metadata_json'])
                    if row['province'] and metadata.get('CD_NO') is not None:
                        parents[(row['province'].strip().upper(), district_number(metadata['CD_NO']))].add(row['label'])
            for row in rows:
                cached = caches.get(row['excel_row'], {})
                resolved = inventory.resolve(row, source['census_vintage'], parents, cached)
                if cached:
                    cache_rows += 1
                    resolved['method'] += '_using_workbook_cached_metadata'
                statuses[resolved['status']] += 1
                endpoint_count += len(resolved['snapshot_ids'])
                if resolved['status'] == 'explicit_identifier_context_conflict':
                    target = inventory.by_id[(resolved['effective_source_code'], source['census_vintage'])]
                    metadata = json.loads(row['source_metadata_json']); metadata.update(cached)
                    differences = []
                    if resolved['effective_province'] != target['province']: differences.append('province')
                    if name_key(resolved['district']) != name_key(target.get('cd_name', '')): differences.append('district')
                    if name_key(target['name']) not in source_names(row['label'], metadata): differences.append('name_or_tier')
                    context_reviews.append(dict(source=key, excel_row=row['excel_row'],
                                                effective_source_code=resolved['effective_source_code'], census_vintage=source['census_vintage'],
                                                source_label=row['label'], gis_label=target['name'],
                                                source_province=row['province'], gis_province=target['province'],
                                                source_district=resolved['district'], gis_district=target.get('cd_name', ''),
                                                differences=';'.join(differences)))
                writer.writerow(dict(source=key, unit_id=row['unit_id'], excel_row=row['excel_row'],
                                     source_code=row['source_code'], label=row['label'], province=row['province'],
                                     census_vintage=source['census_vintage'], reporting_level=row['reporting_level'],
                                     district=resolved['district'], status=resolved['status'], method=resolved['method'],
                                     snapshot_ids_json=json.dumps(resolved['snapshot_ids']),
                                     same_code_snapshot=resolved['same_code_snapshot'], same_code_gis_name=resolved['same_code_gis_name'],
                                     effective_source_code=resolved['effective_source_code'],
                                     effective_province=resolved['effective_province'],
                                     cached_metadata_json=json.dumps(cached, ensure_ascii=False)))
            db.close()
    if sum(statuses.values()) != catalog['reporting_rows']:
        raise ValueError('Source binding assessment did not preserve every reporting row')
    validation = validate_bindings(root, representations, out/'bindings.csv')
    (out/'validation.json').write_text(json.dumps(validation, indent=2)+'\n')
    if validation['errors']:
        raise ValueError(validation)
    with (out/'identifier_context_review.csv').open('w', newline='') as stream:
        fields = ['source','excel_row','effective_source_code','census_vintage','source_label','gis_label',
                  'source_province','gis_province','source_district','gis_district','differences']
        writer = csv.DictWriter(stream, fieldnames=fields); writer.writeheader(); writer.writerows(context_reviews)
    result = dict(reporting_rows=sum(statuses.values()), candidate_endpoints=endpoint_count,
                  statuses=dict(statuses), missing_map_endpoints=0, cached_metadata_rows=cache_rows,
                  context_conflict_fields=dict(Counter(field for row in context_reviews for field in row['differences'].split(';'))),
                  source_database_sha256=source_hashes,
                  inventory_sha256=hashlib.sha256(representations.read_bytes()).hexdigest(),
                  assessor_sha256=hashlib.sha256(Path(__file__).read_bytes()).hexdigest(),
                  identity_asserted=False, population_apportioned=False,
                  scope='Source-to-map evidence assessments; no historical identity or adjusted population inference')
    (out/'manifest.json').write_text(json.dumps(result, indent=2)+'\n')
    return result


if __name__ == '__main__':
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--source-root', type=Path, default=REPO_ROOT/'data_quality/lod_census_sources')
    ap.add_argument('--representations', type=Path, default=REPO_ROOT/'data_quality/lod_identity/representations.csv')
    ap.add_argument('--out', type=Path, default=REPO_ROOT/'data_quality/lod_source_bindings')
    args = ap.parse_args()
    result = build(args.source_root, args.representations, args.out)
    print(json.dumps({k: v for k,v in result.items() if k != 'source_database_sha256'}, indent=2))

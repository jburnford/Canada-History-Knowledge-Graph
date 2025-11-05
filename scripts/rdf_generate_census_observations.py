#!/usr/bin/env python3
"""
Generate CIDOC-CRM RDF measurements (E16_Measurement, E54_Dimension) from census tables.

Links census observations to existing E93_Presence nodes in the spatial/temporal RDF.
Reads historical census tables from ZIP archives or extracted directories.

Usage:
    # Generate observations for 1901 Canada-wide
    python3 scripts/rdf_generate_census_observations.py \
        --year 1901 \
        --census-data 1901Tables.zip \
        --base https://canada.census.example.org/ \
        --out generated/canada/canada_observations_1901.ttl

    # Generate observations for specific province
    python3 scripts/rdf_generate_census_observations.py \
        --year 1911 \
        --census-data 1911Tables \
        --province ON \
        --out generated/canada/ontario_observations_1911.ttl

Supported years: 1891, 1901, 1911
"""

import argparse
import io
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import sys


NS = {
    'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
}


def read_sheet_rows_from_zip(zip_path: Path, xlsx_in_zip: str, sheet_name: str) -> List[List[str]]:
    """Read Excel sheet from ZIP archive"""
    with zipfile.ZipFile(zip_path) as oz:
        xbytes = oz.read(xlsx_in_zip)
    z = zipfile.ZipFile(io.BytesIO(xbytes))

    wb = ET.fromstring(z.read('xl/workbook.xml'))
    name2rid = {s.attrib.get('name'): s.attrib.get('{%s}id' % NS['r']) for s in wb.find('x:sheets', NS)}

    if sheet_name not in name2rid:
        raise SystemExit(f"Sheet '{sheet_name}' not found in {xlsx_in_zip}. Available: {sorted(name2rid)}")

    rid = name2rid[sheet_name]
    rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    id2t = {r.attrib['Id']: r.attrib['Target'] for r in rels}
    ws = ET.fromstring(z.read('xl/' + id2t[rid]))

    # Load shared strings
    sst = {}
    if 'xl/sharedStrings.xml' in z.namelist():
        sst_xml = ET.fromstring(z.read('xl/sharedStrings.xml'))
        for idx, si in enumerate(sst_xml.findall('x:si', NS)):
            t = si.find('x:t', NS)
            if t is not None and t.text is not None:
                sst[idx] = t.text
            else:
                sst[idx] = ''.join((r.find('x:t', NS).text or '') for r in si.findall('x:r', NS))

    # Parse rows
    rows = []
    for r in ws.findall('.//x:sheetData/x:row', NS):
        row = []
        for c in r.findall('x:c', NS):
            t = c.attrib.get('t')
            v = c.find('x:v', NS)
            val = v.text if v is not None else ''
            if t == 's':
                try:
                    val = sst.get(int(val), val)
                except Exception:
                    pass
            row.append(val)
        rows.append(row)
    return rows


def read_sheet_rows_from_dir(data_dir: Path, xlsx_file: str, sheet_name: str) -> List[List[str]]:
    """Read Excel sheet from extracted directory (uses openpyxl if available, falls back to zip method)"""
    xlsx_path = data_dir / xlsx_file

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    # Try openpyxl first (faster and more reliable)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Sheet '{sheet_name}' not found in {xlsx_file}. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else '' for cell in row])
        wb.close()
        return rows
    except ImportError:
        # Fall back to XML parsing
        pass

    # Read as ZIP file
    z = zipfile.ZipFile(xlsx_path)
    wb = ET.fromstring(z.read('xl/workbook.xml'))
    name2rid = {s.attrib.get('name'): s.attrib.get('{%s}id' % NS['r']) for s in wb.find('x:sheets', NS)}

    if sheet_name not in name2rid:
        raise SystemExit(f"Sheet '{sheet_name}' not found in {xlsx_file}. Available: {sorted(name2rid)}")

    rid = name2rid[sheet_name]
    rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    id2t = {r.attrib['Id']: r.attrib['Target'] for r in rels}
    ws = ET.fromstring(z.read('xl/' + id2t[rid]))

    sst = {}
    if 'xl/sharedStrings.xml' in z.namelist():
        sst_xml = ET.fromstring(z.read('xl/sharedStrings.xml'))
        for idx, si in enumerate(sst_xml.findall('x:si', NS)):
            t = si.find('x:t', NS)
            if t is not None and t.text is not None:
                sst[idx] = t.text
            else:
                sst[idx] = ''.join((r.find('x:t', NS).text or '') for r in si.findall('x:r', NS))

    rows = []
    for r in ws.findall('.//x:sheetData/x:row', NS):
        row = []
        for c in r.findall('x:c', NS):
            t = c.attrib.get('t')
            v = c.find('x:v', NS)
            val = v.text if v is not None else ''
            if t == 's':
                try:
                    val = sst.get(int(val), val)
                except Exception:
                    pass
            row.append(val)
        rows.append(row)
    return rows


def ttl_escape(s: str) -> str:
    """Escape string for Turtle literals"""
    if not s:
        return ""
    return s.replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n').replace('\r', '\\r')


def infer_unit(col: str, label: str) -> str:
    """Infer measurement unit from column name and label"""
    c = col.upper()
    L = (label or '').upper()

    if 'POP' in c or 'POPULATION' in L:
        return 'person'
    elif '_AC' in c or 'ACRE' in c or 'ACRE' in L:
        return 'acre'
    elif 'SQ_MI' in c or 'SQUARE MILE' in L:
        return 'square_mile'
    elif c.endswith('_BU') or 'BUSHEL' in L:
        return 'bushel'
    elif c.endswith('_LB') or 'POUND' in L:
        return 'pound'
    elif 'TON' in c or 'TON' in L:
        return 'ton'
    elif 'HEAD' in L or ('CATTLE' in L or 'SHEEP' in L or 'SWINE' in L or 'COW' in L):
        return 'head'
    elif 'HOUSE' in c or 'FAMIL' in c:
        return 'count'
    else:
        return 'count'


def build_dataset_configs(year: str) -> List[Dict]:
    """Build dataset configurations for the given year"""
    y = str(year)

    if y == '1901':
        return [
            {
                'xlsx': '1901/1901_V1T7_PUB_202306.xlsx',
                'sheet': 'CA_V1T7_1901',
                'id_col': 'V1T7_1901',
                'var_sheet': 'Variables',
                'table_name': 'V1T7'
            }
        ]
    elif y == '1891':
        return [
            {
                'xlsx': '1891/1891_V1T3_PUB_202306.xlsx',
                'sheet': 'CA_V1T3_1891',
                'id_col': 'V1T3_1891',
                'var_sheet': 'Variables',
                'table_name': 'V1T3'
            },
            {
                'xlsx': '1891/1891_V4T2_PUB_202306.xlsx',
                'sheet': 'CA_V4T2_1891',
                'id_col': 'V4T2_1891',
                'var_sheet': 'Variables',
                'table_name': 'V4T2'
            },
            {
                'xlsx': '1891/1891_V4T3_PUB_202306.xlsx',
                'sheet': 'CA_V4T3_1891',
                'id_col': 'V4T3_1891',
                'var_sheet': 'Variables',
                'table_name': 'V4T3'
            }
        ]
    elif y == '1911':
        return [
            {
                'xlsx': '1911/1911_V1T1_PUB_202306.xlsx',
                'sheet': 'CA_V1T1_1911',
                'id_col': 'V1T1_1911',
                'var_sheet': 'Variables',
                'table_name': 'V1T1'
            },
            {
                'xlsx': '1911/1911_V1T2_PUB_202306.xlsx',
                'sheet': 'CA_V1T2_1911',
                'id_col': 'V1T2_1911',
                'var_sheet': 'Variables',
                'table_name': 'V1T2'
            },
            {
                'xlsx': '1911/1911_V2T2_PUB_202306.xlsx',
                'sheet': 'CA_V2T2_1911',
                'id_col': 'V2T2_1911',
                'var_sheet': 'Variables',
                'table_name': 'V2T2'
            },
            {
                'xlsx': '1911/1911_V2T7_PUB_202306.xlsx',
                'sheet': 'CA_V2T7_1911',
                'id_col': 'V2T7_1911',
                'var_sheet': 'Variables',
                'table_name': 'V2T7'
            },
            {
                'xlsx': '1911/1911_V2T28_PUB_202306.xlsx',
                'sheet': 'CA_V2T28_1911',
                'id_col': 'V2T28_1911',
                'var_sheet': 'Variables',
                'table_name': 'V2T28'
            }
        ]
    else:
        raise SystemExit(f"Year not supported: {year}. Supported: 1891, 1901, 1911")


def tcpuid_from_row_id(row_id: str) -> Optional[str]:
    """Extract TCPUID from row identifier (e.g., 'V1T7_1901' column value)"""
    # Row IDs are typically in format like "ON038001" (TCPUID) or hierarchical IDs
    # We need the TCPUID to match with E93_Presence nodes
    if not row_id:
        return None

    # Remove any table prefixes
    row_id = row_id.strip()

    # TCPUID format: 2-letter province + 3-digit CD + 3-digit CSD
    if len(row_id) >= 8 and row_id[:2].isalpha() and row_id[2:8].isdigit():
        return row_id[:8]

    return None


class ObservationsRDFGenerator:
    """Generate CIDOC-CRM observation RDF from census tables"""

    def __init__(self, year: str, base_uri: str, province_filter: Optional[str] = None):
        self.year = year
        self.base = base_uri.rstrip('/') + '/'
        self.province_filter = province_filter.upper() if province_filter and province_filter != 'ALL' else None
        self.lines = []
        self.measurement_types = set()
        self.units_declared = set()
        self.stats = {'measurements': 0, 'dimensions': 0, 'csds_with_data': 0}

    def add(self, line: str):
        """Add line to output"""
        self.lines.append(line)

    def generate_prefixes(self):
        """Generate RDF prefixes"""
        self.add("@prefix crm: <http://www.cidoc-crm.org/cidoc-crm/> .")
        self.add("@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .")
        self.add("@prefix xsd: <http://www.w3.org/2001/XMLSchema#> .")
        self.add(f"@prefix ex: <{self.base}> .")
        self.add("")
        self.add(f"# Census Observations for {self.year}")
        if self.province_filter:
            self.add(f"# Province filter: {self.province_filter}")
        self.add("")

    def declare_unit(self, unit: str):
        """Declare a measurement unit if not already declared"""
        if unit in self.units_declared:
            return

        self.add(f'ex:unit/{unit} a crm:E58_Measurement_Unit ; rdfs:label "{unit}" .')
        self.units_declared.add(unit)

    def declare_measurement_type(self, code: str, label: str):
        """Declare a measurement type if not already declared"""
        key = f"{code}_{self.year}"
        if key in self.measurement_types:
            return

        self.add(f'ex:measurementType/{key} a crm:E55_Type ; rdfs:label "{ttl_escape(label)} ({self.year})" .')
        self.measurement_types.add(key)

    def process_dataset(self, config: Dict, rows: List[List[str]]):
        """Process a single census table dataset"""
        if not rows:
            return

        hdr = rows[0]
        idx = {h: i for i, h in enumerate(hdr)}

        # Check required columns
        required = [config['id_col'], 'PR', 'CSD_NO']
        for c in required:
            if c not in idx:
                print(f"WARNING: Missing required column {c} in {config['xlsx']}", file=sys.stderr)
                return

        # Load variable labels from Variables sheet if available
        var_labels = {}
        # Note: Variable sheet reading would go here if needed

        # Auto-detect measurement columns
        skip = {config['id_col'], 'PR_CD_CSD', 'ROW_ID', 'PR', 'CD_NO', 'CSD_NO', 'CSD_TYPE', 'NOTES'}
        measures = {}

        for col in hdr:
            if col in skip:
                continue

            # Check if column has numeric data
            col_idx = idx.get(col)
            if col_idx is None:
                continue

            has_numeric = False
            for r in rows[1:50]:  # Sample first 50 rows
                if len(r) <= col_idx:
                    continue
                val = (r[col_idx] or '').strip()
                if not val:
                    continue
                vtxt = val.replace(',', '').replace(' ', '')
                try:
                    float(vtxt)
                    has_numeric = True
                    break
                except:
                    pass

            if not has_numeric:
                continue

            label = var_labels.get(col, col)
            code = col.lower()
            unit = infer_unit(col, label)
            measures[col] = (code, label, unit)

        print(f"  Found {len(measures)} measurement columns in {config['table_name']}", file=sys.stderr)

        # Declare units and measurement types
        for col, (code, label, unit) in measures.items():
            self.declare_unit(unit)
            self.declare_measurement_type(code, label)

        self.add("")
        self.add(f"# Measurements from {config['table_name']}")

        # Process data rows
        csds_processed = set()

        for r in rows[1:]:
            if len(r) < len(hdr):
                continue

            pr = r[idx['PR']] if 'PR' in idx else ''
            csd_no = r[idx['CSD_NO']] if 'CSD_NO' in idx else ''

            # Filter by province
            if self.province_filter and pr != self.province_filter:
                continue

            # Skip non-CSD rows (aggregates)
            if csd_no == '0' or not csd_no:
                continue

            row_id = r[idx[config['id_col']]]
            tcpuid = tcpuid_from_row_id(row_id)

            if not tcpuid:
                continue

            csds_processed.add(tcpuid)
            presence_id = f"{tcpuid}_{self.year}"

            # Generate measurements for available numeric values
            for col, (code, label, unit) in measures.items():
                col_i = idx.get(col)
                if col_i is None or len(r) <= col_i:
                    continue

                val = (r[col_i] or '').strip()
                if not val:
                    continue

                # Parse numeric value
                vtxt = val.replace(',', '').replace(' ', '')
                try:
                    float(vtxt)
                except:
                    continue

                # Generate E16_Measurement and E54_Dimension
                meas_id = f"{presence_id}_{code}"
                dim_id = f"{meas_id}_dim"

                self.add(f'ex:measurement/{meas_id} a crm:E16_Measurement ;')
                self.add(f'  crm:P39_measured ex:presence/{presence_id} ;')
                self.add(f'  crm:P2_has_type ex:measurementType/{code}_{self.year} ;')
                self.add(f'  crm:P40_observed_dimension ex:dimension/{dim_id} ;')
                self.add(f'  crm:P4_has_time-span ex:period/CENSUS_{self.year} .')

                # Determine datatype
                lit_type = 'xsd:integer'
                if '.' in vtxt or 'e' in vtxt.lower():
                    lit_type = 'xsd:decimal'

                self.add(f'ex:dimension/{dim_id} a crm:E54_Dimension ;')
                self.add(f'  crm:P91_has_unit ex:unit/{unit} ;')
                self.add(f'  crm:P90_has_value "{vtxt}"^^{lit_type} .')

                self.stats['measurements'] += 1
                self.stats['dimensions'] += 1

        self.stats['csds_with_data'] = len(csds_processed)
        self.add("")

    def generate(self, data_path: Path) -> str:
        """Generate complete RDF document"""
        self.generate_prefixes()

        # Process all datasets for this year
        configs = build_dataset_configs(self.year)
        is_zip = data_path.suffix == '.zip'

        for config in configs:
            print(f"Processing {config['table_name']}...", file=sys.stderr)

            try:
                if is_zip:
                    rows = read_sheet_rows_from_zip(data_path, config['xlsx'], config['sheet'])
                else:
                    rows = read_sheet_rows_from_dir(data_path, config['xlsx'], config['sheet'])

                self.process_dataset(config, rows)
            except Exception as e:
                print(f"ERROR processing {config['xlsx']}: {e}", file=sys.stderr)
                continue

        return '\n'.join(self.lines) + '\n'


def main():
    parser = argparse.ArgumentParser(
        description='Generate CIDOC-CRM observation RDF from census tables'
    )
    parser.add_argument('--year', required=True, help='Census year (1891, 1901, 1911)')
    parser.add_argument('--census-data', type=Path, required=True,
                        help='Path to census data (ZIP file or extracted directory)')
    parser.add_argument('--base', default='https://canada.census.example.org/',
                        help='Base URI for RDF resources')
    parser.add_argument('--province', default='ALL',
                        help='Province code to filter (e.g., ON, QC) or ALL for entire Canada')
    parser.add_argument('--out', type=Path, required=True, help='Output TTL file path')

    args = parser.parse_args()

    if not args.census_data.exists():
        print(f"ERROR: Census data not found: {args.census_data}", file=sys.stderr)
        sys.exit(1)

    generator = ObservationsRDFGenerator(args.year, args.base, args.province)
    ttl_content = generator.generate(args.census_data)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(ttl_content, encoding='utf-8')

    print(f"\nWrote {args.out}", file=sys.stderr)
    print(f"Size: {len(ttl_content):,} bytes", file=sys.stderr)
    print(f"Lines: {len(ttl_content.splitlines()):,}", file=sys.stderr)
    print(f"Measurements: {generator.stats['measurements']:,}", file=sys.stderr)
    print(f"Dimensions: {generator.stats['dimensions']:,}", file=sys.stderr)
    print(f"CSDs with data: {generator.stats['csds_with_data']:,}", file=sys.stderr)


if __name__ == '__main__':
    main()

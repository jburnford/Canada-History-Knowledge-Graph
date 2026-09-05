#!/usr/bin/env python3
"""Preserve every reporting row and statistical cell of 1911 V1T1.

Source table identifiers are scoped separately from GDB identifiers. DLS
descriptions provide survey identifiers, not invented survey geometries.
Previous-year columns retain their own reference year. No aggregates, missing
values, unmatched units, or survey-township rows are silently discarded.
"""

import argparse
import csv
import hashlib
import json
import os
import re
import sqlite3
import tempfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from _config import CONFIG, REPO_ROOT
from _lod_model import number

DLS_RE = re.compile(r'^\s*T\s*(\d+)\s+R\s*(\d+)\s+M\s*([EW])\s*(\d+)\s*$', re.I)
VARIABLE_UNITS = {'AREA_ACRES': 'acres', 'AREA_SQ_MI': 'square miles',
                  'POP_M': 'persons', 'POP_F': 'persons', 'POP_TOT': 'persons',
                  'POP_PER_SQ_MI': 'persons per square mile', 'POP': 'persons'}
METADATA = {'ROW_ID', 'V1T1_1911', 'PR', 'CD_NO', 'CSD_NO', 'PR_CD_CSD', 'NOTES'}
VALUE_COLUMNS = {'AREA_ACRES_1911', 'AREA_SQ_MI_1911', 'POP_M_1911', 'POP_F_1911',
                 'POP_TOT_1911', 'POP_PER_SQ_MI_1911', 'POP_1901'}


def parse_dls(name):
    match = DLS_RE.fullmatch(name or '')
    if not match:
        return None
    township, range_number, direction, meridian = match.groups()
    if min(int(township), int(range_number), int(meridian)) <= 0:
        return None
    result = dict(township=int(township), range=int(range_number),
                  meridian_direction=direction.upper(), meridian=int(meridian))
    result['survey_unit_id'] = f'DLS_T{result["township"]}_R{result["range"]}_{direction.upper()}{result["meridian"]}'
    return result


def column_definition(header):
    match = re.fullmatch(r'(.+)_(\d{4})', header)
    if not match or match.group(1) not in VARIABLE_UNITS:
        raise ValueError(f'Unrecognized statistical column: {header}')
    variable, year = match.groups()
    return variable, int(year), VARIABLE_UNITS[variable]


def reporting_level(metadata, dls):
    def zero(value):
        return value is not None and str(value).strip() in {'0', '0.0'}
    if str(metadata.get('PR', '')).strip() == 'CA':
        return 'country_total'
    if zero(metadata.get('CD_NO')):
        return 'province_or_territory_total'
    if zero(metadata.get('CSD_NO')):
        return 'census_division_total'
    if dls:
        return 'survey_township_reporting_unit'
    if metadata.get('CSD_NO') is None:
        return 'unresolved_reporting_level'
    return 'census_subdivision_reporting_unit'


def cell_value(raw, data_type):
    if raw is None:
        return None, 'source_blank'
    if data_type == 'f':
        return None, 'source_formula_requires_evaluation'
    value = number(raw)
    return (str(value), 'numeric') if value is not None else (None, 'source_text_requires_interpretation')


def read_mapping(path, key):
    with path.open() as stream:
        rows = list(csv.DictReader(stream))
    mapping = {}
    for row in rows:
        if row[key] in mapping:
            raise ValueError(f'Duplicate mapping key {row[key]} in {path}')
        mapping[row[key]] = row
    return mapping


def stage(workbook_path, crosswalk_path, inventory_path, out):
    out.mkdir(parents=True, exist_ok=True)
    crosswalk = read_mapping(crosswalk_path, 'v1t1_code')
    gis = read_mapping(inventory_path, 'tcpuid')
    source_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet = workbook.worksheets[0]
    iterator = sheet.iter_rows()
    headers = [cell.value for cell in next(iterator)]
    if len(set(headers)) != len(headers) or set(headers) != METADATA | VALUE_COLUMNS:
        raise ValueError('Unexpected or duplicate 1911 V1T1 headers')
    if not METADATA <= set(headers):
        raise ValueError('Required source metadata columns are missing')
    columns = [(i, header, column_definition(header)) for i, header in enumerate(headers) if header not in METADATA]
    fd, temporary = tempfile.mkstemp(prefix='reporting-', suffix='.sqlite', dir=out)
    os.close(fd)
    connection = sqlite3.connect(temporary)
    connection.executescript('''
        CREATE TABLE reporting_units (
            unit_id TEXT PRIMARY KEY, source_code TEXT, excel_row INTEGER UNIQUE,
            label TEXT, province TEXT, reporting_level TEXT, survey_unit_id TEXT,
            candidate_snapshot_id TEXT, spatial_binding_status TEXT,
            legacy_same_code_gis_name TEXT, code_collision INTEGER,
            source_metadata_json TEXT NOT NULL);
        CREATE TABLE observations (
            observation_id TEXT PRIMARY KEY, unit_id TEXT NOT NULL REFERENCES reporting_units(unit_id),
            source_cell TEXT UNIQUE, source_column TEXT, variable TEXT, reference_year INTEGER,
            reporting_geography_vintage INTEGER, unit TEXT, raw_value_json TEXT,
            numeric_value TEXT, value_status TEXT, excel_number_format TEXT);
        CREATE TABLE survey_units (
            survey_unit_id TEXT PRIMARY KEY, township INTEGER, range_number INTEGER,
            meridian_direction TEXT, meridian INTEGER, geometry_status TEXT);
        CREATE TABLE source (path TEXT, sha256 TEXT, sheet TEXT, census_vintage INTEGER);
    ''')
    connection.execute('PRAGMA foreign_keys = ON')
    connection.execute('INSERT INTO source VALUES (?,?,?,?)', (str(workbook_path), source_hash, sheet.title, 1911))
    levels, statuses, cells, bindings = Counter(), Counter(), Counter(), Counter()
    collisions, source_codes, blank_rows = 0, Counter(), 0
    for row_no, row in enumerate(iterator, start=2):
        if all(cell.value is None for cell in row):
            blank_rows += 1
            continue
        values = {header: row[i].value for i, header in enumerate(headers)}
        metadata = {h: values[h] for h in headers if h in METADATA}
        name = str(values['PR_CD_CSD'] or '').strip()
        code = str(values['V1T1_1911'] or '').strip()
        source_codes[code] += 1
        if code and source_codes[code] > 1:
            raise ValueError(f'Duplicate source reporting code needs explicit disambiguation: {code}')
        unit_id = f'REPORTING_1911_V1T1_{code}' if code else f'REPORTING_1911_V1T1_ROW_{row_no}'
        dls = parse_dls(name)
        level = reporting_level(metadata, dls)
        candidate, binding = '', 'no_validated_spatial_binding'
        if dls:
            binding = 'survey_identifier_only_geometry_unresolved'
            connection.execute('INSERT OR IGNORE INTO survey_units VALUES (?,?,?,?,?,?)',
                               (dls['survey_unit_id'], dls['township'], dls['range'],
                                dls['meridian_direction'], dls['meridian'], 'not_supplied'))
        elif level in {'country_total', 'province_or_territory_total', 'census_division_total'}:
            binding = 'aggregate_reporting_unit'
        elif code in crosswalk:
            match = crosswalk[code]
            uid = match['tcpuid']
            if uid not in gis or match['pr'].strip() != str(values['PR']).strip():
                raise ValueError(f'Invalid crosswalk endpoint/province: {code} -> {uid}')
            candidate, binding = f'{uid}_1911', 'legacy_name_match_candidate'
        legacy_name = gis.get(code, {}).get('csd_name', '')
        collision = bool(legacy_name) and (bool(dls) or name.casefold() != legacy_name.casefold())
        collisions += collision
        connection.execute('INSERT INTO reporting_units VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                           (unit_id, code, row_no, name, values['PR'], level,
                            dls['survey_unit_id'] if dls else '', candidate, binding,
                            legacy_name, int(collision), json.dumps(metadata, ensure_ascii=False)))
        levels[level] += 1; bindings[binding] += 1
        for index, header, (variable, reference_year, unit) in columns:
            cell = row[index]
            numeric, status = cell_value(cell.value, cell.data_type)
            coordinate = f'{sheet.title}!{get_column_letter(index+1)}{row_no}'
            connection.execute('INSERT INTO observations VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                               (f'CELL_1911_V1T1_{row_no}_{index+1}', unit_id, coordinate, header, variable,
                                reference_year, 1911, unit, json.dumps(cell.value, ensure_ascii=False),
                                numeric, status, cell.number_format or 'General'))
            statuses[status] += 1; cells[reference_year] += 1
    connection.commit()
    count = connection.execute('SELECT COUNT(*) FROM reporting_units').fetchone()[0]
    cell_count = connection.execute('SELECT COUNT(*) FROM observations').fetchone()[0]
    surveys = connection.execute('SELECT COUNT(*) FROM survey_units').fetchone()[0]
    foreign_key_errors = connection.execute('PRAGMA foreign_key_check').fetchall()
    expected = count * len(columns)
    if cell_count != expected or foreign_key_errors:
        raise ValueError('Source cell reconciliation failed')
    for table in ['reporting_units', 'survey_units']:
        cursor = connection.execute(f'SELECT * FROM {table} ORDER BY 1')
        with (out / f'{table}.csv').open('w') as stream:
            writer = csv.writer(stream); writer.writerow([c[0] for c in cursor.description]); writer.writerows(cursor)
    connection.close(); workbook.close()
    os.replace(temporary, out / 'source_observations.sqlite')
    result = dict(source=str(workbook_path), source_sha256=source_hash,
                  reporting_rows=count, blank_rows=blank_rows, preserved_cells=cell_count,
                  expected_cells=expected, missing_cells=expected-cell_count,
                  survey_units=surveys, reporting_levels=dict(levels), value_statuses=dict(statuses),
                  reference_year_cells=dict(cells), spatial_binding_statuses=dict(bindings),
                  same_code_different_name_review=collisions,
                  duplicate_source_codes=sum(v-1 for v in source_codes.values() if v > 1),
                  foreign_key_errors=foreign_key_errors,
                  source_code_scope='1911 V1T1; never assumed identical to a GDB TCPUID',
                  geographic_comparability_of_previous_year_columns='not_inferred',
                  status='Complete source-cell staging for one table; not national observation export',
                  input_sha256={str(p): hashlib.sha256(p.read_bytes()).hexdigest() for p in
                                [crosswalk_path, inventory_path, Path(__file__), REPO_ROOT / 'scripts/_lod_model.py']})
    (out / 'manifest.json').write_text(json.dumps(result, indent=2) + '\n')
    return result


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--out', type=Path, default=REPO_ROOT / 'data_quality/lod_1911_reporting')
    args = ap.parse_args()
    result = stage(CONFIG.data_root / '1911Tables/1911/1911_V1T1_PUB_202306.xlsx',
                   REPO_ROOT / 'wikidata_grounding/v1t1_1911_crosswalk.csv',
                   REPO_ROOT / 'data_quality/gis_audit_equal_area/csd_inventory_1911.csv', args.out)
    print(json.dumps({k:v for k,v in result.items() if k != 'input_sha256'}, indent=2))


if __name__ == '__main__':
    main()

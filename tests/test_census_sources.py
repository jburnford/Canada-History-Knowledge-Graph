"""Regressions for real workbook ambiguities that previously lost census data."""

import json
import sqlite3
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from rdflib import Graph, Literal, RDF

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / 'scripts'))
from _lod_model import CRM, HGIS
from export_1911_reporting_rdf import QB, export
from stage_census_sources import semantics, stage, structured_sheet, validate_source
from validate_reporting_rdf import validate


def test_units_and_periods_do_not_follow_population_prefix_or_publication_year():
    assert semantics('POP_XF_N', {'description': 'Number of families'}, '1901_V1T7', 1901)['unit'] == 'count'
    dwelling = semantics('T_NUM_OF_DWELLINGS_1921', {'description': 'Number of dwellings'}, '1921_V3T3', 1921)
    assert dwelling['unit'] == 'dwellings'
    assert dwelling['reference_year'] == 1921
    sterling = semantics('GRIST_PL_V', {'description': 'Value of annual production (pounds sterling)'}, '1851_V2T7', 1851)
    assert sterling['unit'] == 'pounds sterling'
    assert sterling['reference_year'] is None
    assert sterling['reference_period_kind'] == 'annual_production_as_described'
    harvest = semantics('WHT_XX_B', {'description': 'Bushels of wheat produced in the past year'}, '1871_V3T23', 1871)
    assert harvest['reference_year'] is None
    assert harvest['reference_period_kind'] == 'preceding_year_as_described'
    unknown = semantics('FRU_AO_V', {'description': 'Value of fruit produced in the past year'}, '1871_PE_V1T1', 1871)
    assert unknown['unit'] is None
    previous = semantics('POP_1901', {'description': 'Population in 1901'}, '1911_V1T1', 1911)
    assert previous['reference_year'] == 1901 and previous['unit'] == 'persons'


def test_structured_sheet_duplicate_headers_and_unknown_cells_survive(tmp_path):
    workbook = Workbook()
    workbook.active.title = 'OCR'
    workbook.active.append(['Table XXII.—Field Products.'])
    sheet = workbook.create_sheet('Structured')
    sheet.append(['YEAR', 'PR', 'NAME_CD_1871', 'NAME_CSD_1871', 'TCPUID_CD_1871',
                  'TCPUID_CSD_1871', 'POP_XX_N', 'POP_XX_N', 'WHT_XX_B', 'FRU_AO_V', None, None])
    sheet.append([1871, 'SK', 'District', 'Township', 'SK001', 'SK001001', 10, 11, 50, '£52 s18 d0', 'unlabelled', None])
    sheet.append([1871, 'SK', 'District', 'Village', 'SK001', 'SK001001', 0, None, '=1+1', 4, None, None])
    path = tmp_path/'1871_V3T23_CSD_TEST.xlsx'
    workbook.save(path)
    master = {k: {'description': v} for k, v in {
        'POP_XX_N': 'Total population', 'WHT_XX_B': 'Bushels of wheat produced in the past year',
        'FRU_AO_V': 'Value of fruit produced in the past year'}.items()}
    out = tmp_path/'out'
    result = stage(path, 1871, out, master)
    assert result['selected_sheet'] == 'Structured'
    assert result['preserved_cells'] == 10
    assert result['duplicate_statistical_headers'] == {'POP_XX_N': 2}
    assert result['duplicate_source_codes'] == {'SK001001': 2}
    assert result['unlabelled_nonempty_columns'] == ['K']
    assert result['ignored_empty_unnamed_columns'] == 1
    database = out/'source_observations.sqlite'
    assert validate_source(database)['errors'] == []
    db = sqlite3.connect(database)
    assert db.execute('SELECT COUNT(DISTINCT unit_id) FROM reporting_units').fetchone()[0] == 2
    assert db.execute("SELECT COUNT(*) FROM observations WHERE source_column='POP_XX_N'").fetchone()[0] == 4
    rdf = out/'source.nt.gz'
    exported = export(database, rdf)
    assert exported['source_cells'] == 10
    assert validate(database, rdf)['errors'] == []
    import gzip
    with gzip.open(rdf, 'rt') as stream:
        graph = Graph().parse(data=stream.read(), format='nt')
    assert len(set(graph.subjects(HGIS.sourceColumn, Literal('POP_XX_N')))) == 2
    assert len(set(graph.subjects(RDF.type, QB.Observation))) == 5
    unknown_cell = next(graph.subjects(HGIS.sourceCell, Literal('Structured!J3')))
    assert graph.value(unknown_cell, HGIS.measurementUnit) is None
    assert graph.value(unknown_cell, HGIS.referenceYear) is None
    assignment = graph.value(unknown_cell, HGIS.documentsAttributeAssignment)
    quantity = graph.value(assignment, CRM.P141_assigned)
    assert graph.value(quantity, CRM.P91_has_unit) is None
    db.execute("UPDATE observations SET raw_value_json='999' WHERE source_cell='Structured!G2'")
    db.commit(); db.close()
    assert any('source statistical cell mismatch' in e for e in validate_source(database)['errors'])


def test_cd_only_publication_and_local_definitions_have_no_invented_csd(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['ROW_ID', 'V3T3_1921', 'PR', 'CD_NO', 'PR_CD', 'T_NUM_OF_DWELLINGS_1921', 'NOTES'])
    sheet.append([1, 'CA000', 'CA', 0, 'Canada', 10, 'Original note'])
    sheet.append([2, 'ON000', 'ON', 0, 'Ontario', 5, None])
    sheet.append([3, 'ON001', 'ON', 1, 'District', 2, None])
    definitions = workbook.create_sheet('Variables')
    definitions.append(['T_NUM_OF_DWELLINGS_1921', 'Number of dwellings'])
    path = tmp_path/'1921_V3T3_PUB_TEST.xlsx'
    workbook.save(path)
    out = tmp_path/'out'
    result = stage(path, 1921, out, {})
    assert result['reporting_levels'] == {'country_total': 1, 'province_or_territory_total': 1, 'census_division_total': 1}
    db = sqlite3.connect(out/'source_observations.sqlite')
    assert db.execute('SELECT DISTINCT unit FROM observations').fetchall() == [('dwellings',)]
    raw = json.loads(db.execute('SELECT raw_values_json FROM definition_rows').fetchone()[0])
    assert raw == ['T_NUM_OF_DWELLINGS_1921', 'Number of dwellings']
    db.close()


def test_multiple_structured_sheets_require_an_explicit_choice():
    workbook = Workbook()
    for sheet in [workbook.active, workbook.create_sheet('Second')]:
        sheet.append(['PR', 'TCPUID_CSD_1871'])
    with pytest.raises(ValueError, match='Expected one structured worksheet'):
        structured_sheet(workbook)


def test_binding_rejects_identifier_conflicts_survey_codes_and_coverage():
    from build_source_spatial_bindings import Inventory
    inventory = Inventory([
        dict(source_id='SK216003', year=1911, province='SK', name='Saskatoon c', cd_name='Saskatoon',
             level='csd', snapshot_id='SK216003_1911', is_coverage_record=False),
        dict(source_id='SK216999', year=1911, province='SK', name='NO DATA', cd_name='Saskatoon',
             level='csd', snapshot_id='SK216999_1911', is_coverage_record=True)])
    row = dict(source_code='SK216003', province='SK', label='T24 R1 MW3',
               reporting_level='survey_township_reporting_unit', survey_unit_id='DLS_T24_R1_W3',
               source_metadata_json=json.dumps({'CD_NO':216}))
    assert inventory.resolve(row, 1911, {})['snapshot_ids'] == []
    row.update(label='Different town', survey_unit_id='', reporting_level='census_subdivision_reporting_unit',
               source_metadata_json=json.dumps({'TCPUID_CSD_1911':'SK216003', 'NAME_CD_1911':'Saskatoon'}))
    assert inventory.resolve(row, 1911, {})['status'] == 'explicit_identifier_context_conflict'
    row.update(label='Saskatoon', source_code='UNRELATED_TABLE_CODE',
               source_metadata_json=json.dumps({'CD_NO':216, 'CSD_TYPE':'C'}))
    resolved = inventory.resolve(row, 1911, {('SK','216'):{'Saskatoon'}})
    assert resolved['snapshot_ids'] == ['SK216003_1911']
    assert resolved['status'] == 'unique_contextual_name_candidate'
    assert inventory.resolve(row, 1911, {('SK','216'):{'Other district'}})['snapshot_ids'] == []
    row.update(label='NO DATA', source_code='SK216999',
               source_metadata_json=json.dumps({'TCPUID_CSD_1911':'SK216999','NAME_CD_1911':'Saskatoon'}))
    assert inventory.resolve(row, 1911, {})['status'] == 'map_identifier_is_coverage_record'


def test_binding_formula_cache_is_explicit_and_dashes_preserve_word_boundaries():
    from build_source_spatial_bindings import Inventory, name_key
    assert name_key('Town-Ville') == name_key('Town—Ville')
    inventory = Inventory([dict(source_id='ON001001', year=1891, province='ON', name='Example, Town—Ville',
                                cd_name='County', level='csd', snapshot_id='ON001001_1891', is_coverage_record=False)])
    row = dict(source_code='=F2&G2', province='ON', label='Example, Town-Ville', survey_unit_id='',
               reporting_level='census_subdivision_reporting_unit',
               source_metadata_json=json.dumps({'TCPUID_CSD_1891':'=F2&G2','NAME_CD_1891':'County'}))
    assert inventory.resolve(row, 1891, {})['status'] == 'source_metadata_formula_without_cached_result'
    assert inventory.resolve(row, 1891, {}, {'TCPUID_CSD_1891':None})['snapshot_ids'] == []
    result = inventory.resolve(row, 1891, {}, {'TCPUID_CSD_1891':'ON001001'})
    assert result['snapshot_ids'] == ['ON001001_1891']
    assert result['effective_source_code'] == 'ON001001'
    row['province'] = 'ON '
    assert inventory.resolve(row, 1891, {}, {'TCPUID_CSD_1891':'ON001001'})['snapshot_ids'] == ['ON001001_1891']


def test_binding_validation_rejects_coverage_and_missing_source_rows(tmp_path):
    import csv
    from build_source_spatial_bindings import build, validate_bindings
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['PR','NAME_CD_1871','NAME_CSD_1871','TCPUID_CSD_1871','POP_XX_N'])
    sheet.append(['ON','County','Township','ON001001',123])
    path = tmp_path/'1871_V1T1_CSD_TEST.xlsx'; workbook.save(path)
    root = tmp_path/'sources'
    result = stage(path, 1871, root/path.stem, {'POP_XX_N':{'description':'Total population'}})
    (root/'catalog.json').write_text(json.dumps({'sources':[result],'reporting_rows':1}))
    representations = tmp_path/'representations.csv'
    representations.write_text('source_id,year,province,name,cd_name,level,snapshot_id,is_coverage_record\n'
                               'ON001001,1871,ON,Township,County,csd,ON001001_1871,False\n'
                               'ON001999,1871,ON,NO DATA,County,csd,ON001999_1871,True\n')
    out = tmp_path/'bindings'
    assert build(root, representations, out)['statuses'] == {'source_identifier_and_context_agree':1}
    bindings = out/'bindings.csv'
    assert not validate_bindings(root, representations, bindings)['errors']
    original = bindings.read_text()
    bindings.write_text(original.replace('ON001001_1871','ON001999_1871'))
    assert any('Coverage record' in error for error in validate_bindings(root, representations, bindings)['errors'])
    bindings.write_text(original.splitlines()[0]+'\n')
    assert 'Source reporting unit sets differ' in validate_bindings(root, representations, bindings)['errors']
